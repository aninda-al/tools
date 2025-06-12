# Copyright (c) Microsoft Corporation.
# Licensed under the MIT License.
# Author: Aninda Pradhan
# Email: v-anipradhan@microsoft.com
# Version: 3.1.2
# Date: 11/15/24

"""
This script does the following.
    1) Collects all the package list from the workbook
    2) Clean up pkg names and delete empty rows
    3) Updates the following columns
        a) M: 3.0-dev Daily Build status   (Y/N)
        b) N: Build status date
        c) O: Version in 2.0
        d) P: Version in 3.0
        e) Q: Latest Version  (Fedora Sources)
        f) R: Fedora Sources Link
        g) T: Upstream Sources Link
        h) D: Need Upgrade (Y/N/Revisit/Remove)
        i) E: Upgrade to version (Version/Revisit/NA)
        j) H: Status (Not Started, Ongoing, PR Raised, PR in Review, Done-Upgrade, Done-FixBuild, Done-OtherChanges, NA-Uptodate)
        k) U: Logs any obvious remarks 
    4) This code tries to find the fedora package information from fedora git if it fails to find the data from packages url 
    5) This code tries to extract the fedora version and release info from 'https://kojipkgs.fedoraproject.org/packages/'
Note: This script does the 1st pass auto analysis to figure out if a package upgraded is needed or not. 
    Any cell color coded other than "White" / labeled "Not_Found" must be reviewed developers.        

Usage:
  python3 pkg_update_analysis-<version>.py

Options:
  GUI based : Please enter paths for workbook, build_state.csv, and SPECS-EXTENDED for 2.0 & 3.0
"""

import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from bs4 import BeautifulSoup
import time
from typing import Any, Callable
from functools import wraps
import requests
import pandas as pd
import datetime
import os
import re
import tkinter as tk
import urllib.request
import sys
from pathlib import Path
from tkinter import messagebox
import subprocess
import logging
# import psutil
import git  # pip install gitpython
from packaging.version import Version
from pyrpm.spec import Spec, replace_macros
from datetime import datetime
from urllib.parse import unquote

# from git import Repo

# workbook = 'C:\\Users\\v-anipradhan\\Downloads\\all-1563-pkg-list3.1.2.xlsx'
# # workbook = 'C:\\Users\\v-anipradhan\\Downloads\\AutoAnalyzed_30_NameswithVersion1.xlsx'
# # workbook = r"C:\Users\v-anipradhan\Downloads\testpkgs-missedpkgs1.xlsx"
# build_state = 'C:\\Users\\v-anipradhan\\Downloads\\pkg-artifacts-102424\\pkg_artifacts\\build_state.csv'

# code_dir_3_0 = 'C:\\Users\\v-anipradhan\\work\\al3_0\\azurelinux'
# code_dir_2_0 = 'C:\\Users\\v-anipradhan\\work\\al2_0\\azurelinux'

workbook = ''
build_state = ''
code_dir_3_0 = ''
code_dir_2_0 = ''

# extract_fedora_info_from_kojipkgs = True
# kojipkgs_url = 'https://kojipkgs.fedoraproject.org/packages/'
extract_fedora_info_from_kojipkgs_pkgid = True
koji_pkgid_url = 'https://koji.fedoraproject.org/koji/packageinfo?packageID='
entries = []
pkg_list = []

revisit_color = "CCD519" #yellow
white_color = "FFFFFF"
orange_color = "FF6433"

cur_stable_fedora_rel = 'fc41'
fedora_git_branch='f41'
special_chars = ",!?~^*%$#@"

tmp_git_dir = 'tmp-git-dir'

# Create and configure logger
logging.basicConfig(filename="pkg_update_analysis.log",
                    format='%(asctime)s [%(filename)s:%(lineno)s - %(funcName)20s() ] %(message)s',
                    filemode='w')
# Creating an object
logger = logging.getLogger()

# Setting the threshold of logger to DEBUG
logger.setLevel(logging.DEBUG)
    
    
def is_file_open(filepath):
    """This method iterates over all running processes on your system and checks if any of them have the file open

    Args:
        filepath (Any): Path of excel file

    Returns:
        bool: return True if file is open, else False
    """
    for proc in psutil.process_iter():
        try:
            for open_file in proc.open_files():
                if open_file.path == filepath:
                    return True
        except psutil.NoSuchProcess:
            pass  # Process might have terminated
    return False

def read_all_pkg_names():
    """Load the workbook
       Access the worksheet (assuming it's the first sheet) 
       Read the values from column 'C'
    """
    logger.info(f'workbook={workbook}')
    wb = openpyxl.load_workbook(workbook)  
    ws = wb.active
    
    for row in ws.iter_rows(min_row=3, values_only=True): #min_row=3 as pkg names start from row 3
        
        if row[2] != None:
            pkg = row[2].strip(' \t\n\r')
            pkg_list.append(pkg) # row[2] = our col 'C'

    logger.info(pkg_list)


def updatexl_pkg_col_value(pkg, col, val, color=white_color):
    """_ Update a particular cell for a given package _

    Args:
        pkg (_string_): _ Name of the package _
        col (_int_): _ column number _
        val (_any_): _ value _
    """
    # Load the workbook
    wb = openpyxl.load_workbook(workbook)  
    # Access the worksheet (assuming it's the first sheet) 
    ws = wb.active
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r[2] != None:
            if r[2].strip(' \t\n\r') == pkg:
                ws.cell(row=r[1]+2, column=col).value = val
                if color!=white_color:
                    ws.cell(row=r[1]+2, column=col).fill = PatternFill("solid", fgColor=color)
                elif color==white_color and val == "Not_Found":
                    ws.cell(row=r[1]+2, column=col).fill = PatternFill("solid", fgColor=orange_color)
                elif color==white_color:
                    ws.cell(row=r[1]+2, column=col).fill = PatternFill("solid", fgColor=white_color)
    wb.save(workbook)    
        

def is_fedora_greater(fedora_ver, ver_3_0):
    """ Compares the fedora and Azurelinux 3.0 version strings

    Args:
        fedora_ver (_type_): _description_
        ver_3_0 (_type_): _description_

    Returns:
        _type_: _description_
    """
    result = False
    # # print(f'fedora_ver:{fedora_ver}, ver_3_0:{ver_3_0}')
    # fed_ver = fedora_ver.split('.')
    # az_ver = ver_3_0.split('.')
    # for f, a in zip(fed_ver, az_ver):
    #     # print(f'f:{f} a:{a}')
    #     if int(f) > int(a):
    #         result = True
    #         break
    # # print(f'result = {result}')
    
    # replace all '_' with '.' before comparing
    fedora_ver = fedora_ver.replace("_", ".")
    ver_3_0 = ver_3_0.replace("_", ".")
    try:
        if Version(fedora_ver) > Version(ver_3_0):
            return True
    except Exception as e:
        logger.exception(f"Error comparing versions: {e} , fedora_ver: {fedora_ver}, ver_3_0: {ver_3_0}") 
        print(f"Error comparing versions: {e} , fedora_ver: {fedora_ver}, ver_3_0: {ver_3_0}")
    return result

def parse_version(version):
    # Replace special cases
    version = version.replace('Not_Found', '0').replace('Not_found', '0').replace('_', '.')
    
    # Split the version into components
    # This regex captures numbers, letters, and special characters
    components = re.split(r'(\d+|[a-zA-Z]+|[^0-9a-zA-Z]+)', version)
    # Filter out empty strings and return a list of components
    return [comp for comp in components if comp]

def is_fedora_version_greater(ver_3_0, fedora_ver):
    # print(f'ver_3_0 : {ver_3_0} , fedora_ver : {fedora_ver}')
    # ver_3_0 = str(ver_3_0)
    # fedora_ver = str(fedora_ver)
    parsed_v1 = parse_version(ver_3_0)
    parsed_v2 = parse_version(fedora_ver)
    # print(f'ver_3_0 parsed : {parsed_v1} , fedora_ver parsed: {parsed_v2}')
    # Compare the parsed components
    for comp1, comp2 in zip(parsed_v1, parsed_v2):
        if str(comp1).isdigit() and str(comp2).isdigit():
            # Compare as integers
            if int(comp1) < int(comp2):
                return True
            elif int(comp1) > int(comp2):
                return False
        else:
            # Compare as strings
            if str(comp1) < str(comp2):
                return True
            elif str(comp1) > str(comp2):
                return False

    # If all compared components are equal, check lengths
    if len(parsed_v1) < len(parsed_v2):
        return True
    elif len(parsed_v1) > len(parsed_v2):
        return False
    
    return False  # They are equal

def update_if_need_upgrade():
    """ Based on the values fedora version, Azure 3.0 version and build status 
        update the status column if an update is required or not
    """

    # Load the workbook
    wb = openpyxl.load_workbook(workbook)  
    # Access the worksheet (assuming it's the first sheet) 
    ws = wb.active
    for pkg in pkg_list:
        need_upgrade = ''
        fedora_ver = ''    
        for r in ws.iter_rows(min_row=3, values_only=True):
            if r[2] != None:
                if r[2].strip(' \t\n\r') == pkg: #Skip row 0 & 1 , pkg info starts from row 2
                    fedora_ver = ws.cell(row=r[1]+2, column=17).value # col 'Q'
                    # print(f'fedora_ver type = {type(fedora_ver)}')
                    ver_3_0 = ws.cell(row=r[1]+2, column=16).value # col 'P'

                    build_status = ws.cell(row=r[1]+2, column=13).value # col 'M'
                    need_upgrade = ''
                    upgrade_to_ver = ''
                    logger.info(f'fedora = {fedora_ver} : Azure3.0 = {ver_3_0}')
                    if fedora_ver == 'Not_Found' or fedora_ver == None or ver_3_0 == None or ver_3_0 == 'Not_Found' or build_status == 'Not_Found' or build_status == None:
                        need_upgrade = 'Revisit' 
                        upgrade_to_ver = 'Revisit' 
                    elif fedora_ver == ver_3_0:
                        logger.info(f"{pkg}:Upgrade NOT Needed")
                        need_upgrade = 'N'
                        upgrade_to_ver = 'NA'
                    # elif is_fedora_greater(fedora_ver, ver_3_0) : 
                    elif is_fedora_version_greater(ver_3_0, fedora_ver):
                        logger.info(f"{pkg}:Upgrade Needed")
                        need_upgrade = 'Y'
                        upgrade_to_ver = fedora_ver

                    # elif build_status == 'Not_Found':
                    #     need_upgrade = 'Revisit'
                    #     upgrade_to_ver = 'Revisit'
                    else: # Dont expect it to ever reach here
                        need_upgrade = 'Revisit'
                        upgrade_to_ver = 'Revisit'                                    

                    logger.info(f'pkg = {pkg} , need_upgrade = {need_upgrade}, fedora_ver = {fedora_ver}')
                    if need_upgrade == 'Y' or need_upgrade == 'N':
                        updatexl_pkg_col_value(pkg, 4, need_upgrade) # col 'D'
                        updatexl_pkg_col_value(pkg, 5, upgrade_to_ver) # col 'E'
                    else:
                        updatexl_pkg_col_value(pkg, 4, need_upgrade, revisit_color) # col 'D' , lets color code revisit
                        updatexl_pkg_col_value(pkg, 5, upgrade_to_ver, revisit_color) # col 'E'


# update col 'M' and 'N' with build_state & Date 
def updatexl_build_status_and_date(pkg, pkg_state):
    """_summary_

    Args:
        pkg (_type_): _description_
        pkg_state (_type_): _description_
    """

    timestamp = os.path.getmtime(build_state)
    date = datetime.fromtimestamp(timestamp).strftime('%m-%d-%Y')
    
    if pkg_state == 'Built':
        pkg_state = 'Y'
    elif pkg_state == 'Failed':
        pkg_state = 'N'
    else:
        pkg_state = 'Not_Found'
        
    logger.info(f'pkg = {pkg} , pkg_state = {pkg_state}')
    updatexl_pkg_col_value(pkg, 13, pkg_state) # col 'M'
    updatexl_pkg_col_value(pkg, 14, date) # col 'N'



# 
def updatexl_fedora_src_link(pkg, version, release):
    """ update col 'R' with Fedora Sources Link 

    Args:
        pkg (_type_): _description_
        version (_type_): _description_
        release (_type_): _description_
    """
    logger.info(f'pkg = {pkg} , version = {version}')
    url = f'https://kojipkgs.fedoraproject.org//packages/{pkg}/{version}/{release}/src/{pkg}-{version}-{release}.src.rpm'
    response = requests.get(url)
    if response.status_code == 200:
        logger.info('src rpm url exists')

        # if cur_stable_fedora_rel not in release:
        #     updatexl_pkg_col_value(pkg, 18, 'Not_Found') # col 'R'
        # else:
        #     updatexl_pkg_col_value(pkg, 18, url) # col 'R'
        updatexl_pkg_col_value(pkg, 18, url) # col 'R'
    else:
        logger.error(f'{url} does not exist')
        url = "Not_Found"
        updatexl_pkg_col_value(pkg, 18, url) # col 'R'   

            
def get_upstream_src(url):
    """_summary_

    Args:
        url (_type_): _description_

    Returns:
        _type_: _description_
    """
    result = "Not_Found"
    source = requests.get(url).text
    soup = BeautifulSoup(source, 'lxml')
    try:
        li_tags = soup.find_all('li')
        # print(f'upstream_src = {upstream_src}')
        for tag in li_tags:
            if "Upstream" in tag.text:
                logger.debug("Found")
                logger.info(tag.a['href'])
                result = tag.a['href']
    except:
        logger.error(f'Error parsing upstream src for url  {url}')
    return result        

    
def get_fedora_pkg_release_info(url):
    """_summary_

    Args:
        url (_type_): _description_

    Returns:
        _type_: _description_
    """
    release = "Not_Found"
    try:
        html_page_table = pd.read_html(url, match="Fedora ")
        df = html_page_table[0] # dataframe of table with all fedora releases and version details
        # print(df)
        for row in df.itertuples():
            logger.info(f'row.Release = {row.Release} row.Stable = {row.Stable}')
            if 'Fedora 41' in row.Release:
                release_str =  row.Stable; 
                release = release_str.split('-') 
                break
    except Exception as e:
        logger.exception(f"Error reading HTML: {e}")  

    logger.info(release)
    return release


def search_in_file(file_path, search_text):
    """
    Searches for a specific text in a file.

    Args:
        file_path (str): The path to the file.
        search_text (str): The text to search for.

    Returns:
        list: A list of lines containing the search text.
    """
    return_line = ''
    with open(file_path, 'r', encoding="utf8") as file:
        for line in file:
            if search_text in line:
                return_line = line
                break

    return  return_line


def get_version_info_from_specfile(file_path):
    """_summary_

    Args:
        file_path (_type_): _description_

    Returns:
        _type_: _description_
    """
    spec = Spec.from_file(file_path)
    if '%' in spec.version:
        version = replace_macros(spec.version, spec)
        if '%{nil}' in version:
            version = version.replace('%{nil}', "")
        if '%(echo' in version or ':Requirements' in version or '%' in version: # some pkg files have run linux commands to resolve macros at runtime perl-Crypt-PasswdMD5.spec, perl-Version-Requirements.spec
            logger.error(f'ERROR: could not resolve pkg Version macro for  {spec.name}')
            version = 'Not_Found'
    else:
        version = spec.version
        
    print(f'{replace_macros(spec.name, spec)} version = {version}')
    return version

def get_release_info_from_specfile(file_path):
    """_summary_

    Args:
        file_path (_type_): _description_

    Returns:
        _type_: _description_
    """
    spec = Spec.from_file(file_path)
    if '%' in spec.release:
        release = replace_macros(spec.release, spec)
        if '%' in release:
            logger.error(f'ERROR: could not resolve pkg Release macro for  {spec.name}')
            release = 'Not_Found'
    else:
        release = spec.release

        
    logger.info(f'{replace_macros(spec.name, spec)} release = {release}')
    # rel_line = search_in_file(file_path, 'Release:')
    # rel_line = rel_line.strip() #strip \n\r
    # release = ''
    # p = r'[\d]+[.,\d]+|[\d]*[.][\d]+|[\d]+' # some spec files used \t , so had to go this route to search numbers
    # release = re.findall(p,  rel_line) # returns list of number string
    # print(f'release from spec {release}')
    # if release != []: # some specs have no release , eg. libtomath
    #     release = release[0]
    return release

def find_file(filename, directory):
    """Finds a specific file in a directory and its subdirectories.

    Args:
        filename (_type_): _description_
        directory (_type_): _description_

    Returns:
        string: file path
    """

    for root, dirs, files in os.walk(directory):
        if filename in files:
            return os.path.join(root, filename)
    return None


def get_pkg_ver(pkg, pkg_ver):
    """ For both 2.0 & 3.0 
        searches the package spec file in SPECS-EXTENDED and if not found searches for the same in SPEC dir
        Extracts the package version details from the SPEC files

    Args:
        pkg (_type_): _description_
        pkg_ver (_type_): _description_

    Returns:
        _type_: _description_
    """
    pkg_ver_info ='Not_Found'
    filename = f'{pkg}.spec'
    if pkg_ver == '3_0':
        directory = code_dir_3_0
    elif pkg_ver == '2_0':
        directory = code_dir_2_0
    else:
        logger.error(f'ERROR: wrong {pkg_ver}')
        return None
    file_path = find_file(filename, directory)
    if file_path:
        logger.info(f"File found: {file_path}")
        if "SPECS-EXTENDED" in file_path:
            pkg_ver_info = get_version_info_from_specfile(file_path)
        elif "SPECS" in file_path:
            pkg_ver_info = get_version_info_from_specfile(file_path)
            pkg_ver_info = f"Moved_to_Core: {pkg} version {pkg_ver_info}"
    else:
        logger.error("File not found.")
    logger.info(f'pkg = {pkg} , pkg_ver = {pkg_ver}, pkg_ver_info = {pkg_ver_info}')
    
    return pkg_ver_info
       
def discard_after_special_chars(string, special_chars):
    """Discards the part of a string after encountering any of the specified special characters.

    Args:
        string (_type_): _description_
        special_chars (_type_): _description_

    Returns:
        _type_: _description_
    """
    for i, char in enumerate(string):
        if char in special_chars:
            return string[:i]
    return string

def get_fedora_pkg_info_from_pkg_url(pkg):
    """ Extract the pkg info for a given package from fedora pkg url 'https://packages.fedoraproject.org/pkgs/<pkg_name>/<pkg_name>'

    Args:
        pkg (string): package name

    Returns:
        tuple: fedora_version, fedora_release_str, upstream_src
    """
    fedora_version = 'Not_Found'
    fedora_release_str = 'Not_Found'
    upstream_src = 'Not_Found'
    url = f'https://packages.fedoraproject.org/pkgs/{pkg}/{pkg}/'
        #check if url exists
    response = requests.get(url)
    if response.status_code == 200:
        logger.info(f'{url} exists')
        try:
            release_info = get_fedora_pkg_release_info(url) 
            if release_info != 'Not_Found':
                fedora_version = release_info[0]  # pkg version used by fedora
                fedora_release_str = release_info[1]  # pkg release info specific to fedora  
            
            upstream_src = get_upstream_src(url)  
        except:
            logger.error(f'url exist for pkg {pkg}, but could not be parsed')     
    else:
        logger.error(f'url does not exist for pkg {pkg}')
    return (fedora_version, fedora_release_str, upstream_src)

def get_fedora_git_branch_name(git_url):
    """_summary_

    Args:
        url (_type_): _description_

    Returns:
        _type_: _description_
    """
    branch_name = ''
    g = git.cmd.Git()
    branches = g.ls_remote("--heads", git_url).split('\n')
    for  branch_str in reversed(branches):
        branch_str = branch_str.strip() #strip \n\r
        branch_token = branch_str.split('/')
        branch_name = branch_token[-1]
        # libraw has branch name 'pkgconfig-private'
        # some branch name has 'fc' like libuninameslist has branch named 'fc6'
        # In reverse order some git branches are listed as for ex in libthai ['refs/heads/rawhide', 'refs/heads/main', 'refs/heads/f9', 'refs/heads/f41']
        if len(branch_name) == 3 and 'fc' not in branch_name: 
            print(branch_name)
            break
    return branch_name

def get_commit_id(clone_path):
    repo = git.Repo(clone_path)
    # Get the commit log
    commits = list(repo.iter_commits())
    # Display short log for the last 5 commits
    for commit in commits[:5]:
        if "Fedora_41_Mass_Rebuild" in commit.message.splitlines()[0]:
            return commit.hexsha[:7]
    return None

def update_missing_macros(file_path, url):
    url = re.split('%|{|}', url)
    print(url[-2])
    ret_str = None
    if 'name' == url[-2]:
        # url = get_macro_value(file_path, 'Name')
        ret_str = search_in_file(file_path, 'Name')
        
    else:
        ret_str = search_in_file(file_path, url[-2])
        
    if ret_str != None:
        ret_str = ret_str.strip() #strip \n\r
        ret_str = re.split(' |\t', ret_str)
        ret_str = ret_str[-1]
    url = url[0] + ret_str
    return url


def get_upstream_url_from_specfile(file_path):
    """_summary_

    Args:
        file_path (_type_): _description_

    Returns:
        _type_: _description_
    """
    url = 'Not_Found'
    spec = Spec.from_file(file_path)
    if spec.url != None:
        if '%' in spec.url:
            url = replace_macros(spec.url, spec)
            if '%' in url: # handle unresolved macros
                logger.error(f'ERROR: could not resolve pkg URL macro for  {spec.name}')
                url = 'Not_Found'
        else:
            url = spec.url

        
    logger.info(f'{replace_macros(spec.name, spec)} URL = {url}')
    # url_line = search_in_file(file_path, 'URL:')
    # url_line = url_line.strip() #strip \n\r   
    # url = re.split(' |\t', url_line)
    # url = url[-1]
    # if '%' in url:
    #     url = update_missing_macros(file_path, url)
    return url
    
    
def get_fedora_pkg_info_from_git(pkg):
    """ Extract the pkg info for a given package from fedora git "https://src.fedoraproject.org/rpms/<pkg_name>.git"

    Args:
        pkg (string): package name

    Returns:
        tuple: fedora_version, fedora_release_str, upstream_src
    """
    fedora_version = 'Not_Found'
    fedora_release_str = 'Not_Found'
    upstream_src = 'Not_Found'
    try:
        if not os.path.exists(tmp_git_dir):
            os.mkdir(tmp_git_dir, exist_ok=True)
            print("Folder created successfully!")
        
        git_url = f'https://src.fedoraproject.org/rpms/{pkg}.git'
        #check if url exists
        try:
            response = requests.get(git_url)
            if response.status_code == 200:
                print(f"{git_url} is valid and exists on the internet")
                clone_path = f'./{tmp_git_dir}/{pkg}'
                fedora_git_branch = get_fedora_git_branch_name(git_url) # We need to find this for every pkg, as this is needed to clone the right git branch and also use it construct the .src.rpm url
                if not os.path.exists(clone_path): # skip if already checked out                    
                    repo = git.Repo.clone_from(git_url, clone_path, branch=fedora_git_branch)
                    # disabling the below code will use the refs/head/f41 
                    # commit_id = get_commit_id(clone_path)
                    # print(f"pkg:{pkg} fedora_git_branch:{fedora_git_branch} commit_id:{commit_id}")
                    # if commit_id != None:
                    #     repo.git.checkout(commit_id)
                        
                    
                #check if SPEC file for the pkg exists
                filename = f'{pkg}.spec'
                file_path = find_file(filename, tmp_git_dir)
                if file_path != None: # to handle pks like libwpe, which has no spec file
                    fedora_version = get_version_info_from_specfile(file_path)
                    fedora_release = get_release_info_from_specfile(file_path)
                    upstream_src = get_upstream_url_from_specfile(file_path)
                    print(f'fedora_version: {fedora_version}  fedora_release: {fedora_release}  upstream_src: {upstream_src}')
                    if fedora_release != None: # Construct fedora release string, needed to compose .src.rpm path 
                        #insert 'c' after 'f' , ex. f41 -> fc41
                        fedora_git_branch = fedora_git_branch[:1] + 'c' + fedora_git_branch[1:]
                        fedora_release_str = f'{fedora_release}.{fedora_git_branch}'
        except requests.ConnectionError as exception:
            print(f"{git_url} does not exist on Internet: exception : {exception}")
            
    except OSError as e:
        print(f"An error occurred: {e}")
        
    return (fedora_version, fedora_release_str, upstream_src)


def get_latest_fedora_release(pkg_url):
    """Check if the link's href ends with .fc41/.
    Extract the date next to the link and determine the latest date.

    Args:
        pkg_url (_type_): _description_

    Returns:
        _type_: _description_
    """
    latest_fedora_rel = 'Not_Found'

    response = requests.get(pkg_url)
    if response.status_code == 200:
        print(f"Pkg URL {pkg_url}: EXIST")

        soup = BeautifulSoup(response.content, 'html.parser')

        # Find all 'a' tags
        links = soup.find_all('a')

        # Initialize variables to track the latest date and href
        latest_date = None
        latest_href = None

        # Regular expression to extract dates in the format 'YYYY-MM-DD HH:MM'
        date_pattern = re.compile(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}')

        # Iterate over all links to find the latest one with trailing '.fc41/'
        for link in links:
            href = link.get('href')
            if href and href.endswith('.fc41/'):
                # Find the date text corresponding to the link
                date_text = link.find_next_sibling(string=True)
                if date_text and date_pattern.search(date_text):
                    date_str = date_pattern.search(date_text).group()
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d %H:%M')
                    if latest_date is None or date_obj > latest_date:
                        latest_date = date_obj
                        latest_href = href
        if latest_href != None:                 
            latest_fedora_rel = latest_href[:-1]
        logger.info(f"Latest href with trailing '.fc41/':{latest_href}")
    else:
        logger.error(f"ERROR: Pkg URL {pkg_url}: Does Not EXIST")
        print(f"ERROR: Pkg URL {pkg_url}: Does Not EXIST")
    return latest_fedora_rel   
        
# Function to parse and split version strings
def parse_version(v):
    # Replace invalid characters with dots for splitting
    cleaned = re.sub(r'[^0-9a-zA-Z]', '.', v)
    return [int(x) if x.isdigit() else x for x in cleaned.split('.')]
    

def get_version_list_from_pkg_url(pkg_url):
    response = requests.get(pkg_url)
    if response.status_code == 200:
        logger.info(f"Pkg URL {pkg_url}: EXIST")
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(response.content, 'html.parser')

        # Find all 'a' tags and get their href values
        links = soup.find_all('a')

        # Extract, decode, and filter version values
        version_list = []
        for link in links:
            href = link.get('href')
            if href and href.endswith('/'):
                # Strip leading/trailing slashes and decode URL-encoded characters
                version = unquote(href.strip('/'))
                # if '~b' in version or 'E' in version or 'b' in version: # Dangerous way to handle 'PyGreSQL'  ver '6.0~b1 and Xaw3d ver '1.5E' , stunnel ver '5.05b5'
                #     version = re.sub(r'~b|[A-Eb]', '.0', version)
                # Filter out any entries containing 'packages'
                if 'packages' not in version:
                    # if 'ii' not in version:
                    #     if '+dev' in version:
                    #         version = version.replace('+dev', '.0')#re.sub(r'+dev', '.0', version)
                    #     if 'alpha' in version:
                    #         version = version.replace('alpha', '')#re.sub(r'+dev', '.0', version)
                    #     if 'git' not in version:
                    #         version = re.sub(r'~alpha|eta|svn|~b|[A-Ea-z]', '.0', version)
                    version_list.append(version)

        logger.info("List of versions:")
        logger.info(version_list)
        # for version in version_list:
        #     print(version)
        return version_list

def get_sorted_version_list(pkg_url):
    version_list = get_version_list_from_pkg_url(pkg_url)
    # Sort the list using the custom sorting function
    sorted_versions = []
    if version_list != []:
        try:
            sorted_versions = sorted(version_list, key=parse_version, reverse=True)
        except:
            logger.error("ERROR sorting pkg url {pkg_url}")
            return []
    else:
        return []
    
    logger.info("Sorted version values:")
    logger.info(sorted_versions)
    # for v in sorted_versions:
    #     print(v)
    return sorted_versions   

def get_fedora_release_info(pkg_url, search_str):
    """Check if the link's href ends with given search string.
    Extract the date next to the link and determine the latest date.

    Args:
        pkg_url (_type_): _description_

    Returns:
        _type_: _description_
    """
    fedora_rel = 'Not_Found'

    response = requests.get(pkg_url)
    if response.status_code == 200:
        logger.info(f"Pkg URL {pkg_url}: EXIST")

        soup = BeautifulSoup(response.content, 'html.parser')

        # Find all 'a' tags
        links = soup.find_all('a')

        # Initialize variables to track the latest date and href
        latest_date = None
        latest_href = None

        # Regular expression to extract dates in the format 'YYYY-MM-DD HH:MM'
        date_pattern = re.compile(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}')

        # Iterate over all links to find the latest one with trailing '.fc41/'
        for link in links:
            href = link.get('href')
            if href and href.endswith(search_str):
                # Find the date text corresponding to the link
                date_text = link.find_next_sibling(string=True)
                if date_text and date_pattern.search(date_text):
                    date_str = date_pattern.search(date_text).group()
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d %H:%M')
                    if latest_date is None or date_obj > latest_date:
                        latest_date = date_obj
                        latest_href = href
        if latest_href != None:                 
            fedora_rel = latest_href[:-1]
        logger.info(f"Latest href with trailing '{search_str}':{latest_href}")
    else:
        logger.error(f"ERROR: Pkg URL {pkg_url}: Does Not EXIST")
        print(f"ERROR: Pkg URL {pkg_url}: Does Not EXIST")
    return fedora_rel


def get_latest_href_from_pkg_url(pkg_url):
    '''
    For a given url finds the href attributes and their corresponding dates. 
    It compares the dates to find the href with the latest date.
    '''
    # latest_fedora_ver = 'Not_Found'
    # Send a GET request to the URL
    response = requests.get(pkg_url)
    if response.status_code == 200:
        logger.info(f"Pkg URL {pkg_url}: EXIST")
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(response.content, 'html.parser')

        # Find all 'a' tags and their corresponding text
        links = soup.find_all('a')

        # Initialize variables to track the latest date and href
        latest_date = None
        latest_href = 'Not_Found'

        # Regular expression to extract dates in the format 'YYYY-MM-DD HH:MM'
        date_pattern = re.compile(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}')

        # Iterate over all links to find the latest one
        for link in links:
            # Find the href attribute
            href = link.get('href')
            if href:
                # Find the date text corresponding to the link
                date_text = link.find_next_sibling(string=True)
                if date_text and date_pattern.search(date_text):
                    date_str = date_pattern.search(date_text).group()
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d %H:%M')
                    if latest_date is None or date_obj > latest_date:
                        latest_date = date_obj
                        latest_href = href
        if latest_href != 'Not_Found':
            latest_href = latest_href[:-1]
        logger.info(f"Latest href: {latest_href}")
    else:
        logger.error(f"ERROR: Pkg URL {pkg_url}: Does Not EXIST")
        print(f"ERROR: Pkg URL {pkg_url}: Does Not EXIST")
    return latest_href


def get_fedora_pkg_info_from_kojipkgs_url(pkg):
    """ Extracts fedora pkg info from kojipkgs url

    Args:
        pkg (_type_): _description_

    Returns:
        _type_: _description_
    """
    pkg_url = f'{kojipkgs_url}{pkg}'
    sorted_versions = get_sorted_version_list(pkg_url)
    fedora_version = 'Not_Found'
    fedora_release = 'Not_Found'
    if sorted_versions != []:
        fedora_rel_info = get_fedora_release_info(f'{pkg_url}/{sorted_versions[0]}', f'.fc41/')
        if 'fc41'in fedora_rel_info: # check if sorted_versions[0] has fc41
            fedora_version = sorted_versions[0]
            fedora_release = fedora_rel_info
            return fedora_version, fedora_release
        elif fedora_rel_info == 'Not_Found' or 'fc42' in get_fedora_release_info(f'{pkg_url}/{sorted_versions[0]}', f'.fc42/'): #sorted_versions[0] has fc42 but no fc41 or handle empty version dir like 'cairomm'
            # loop through the sorted_versions to find a fc41 release 
            if len(sorted_versions) > 1: # there are more than one fedora versions, lets loop through each to find fc41
                i = 1 # start from 1 since 0th one already taken care during the very first search
                for i in range(len(sorted_versions)):
                    print(f'searching fedora version : {sorted_versions[i]} for release fc41')
                    fedora_rel_info = get_fedora_release_info(f'{pkg_url}/{sorted_versions[i]}', f'.fc41/')
                    if 'fc41'in fedora_rel_info: # check if sorted_versions[i] has fc41
                        fedora_version = sorted_versions[i]
                        fedora_release = fedora_rel_info
                        return fedora_version, fedora_release
                    else:
                        i = i + 1
            else: #
                fedora_rel_info = get_latest_href_from_pkg_url(f'{pkg_url}/{sorted_versions[0]}')
                fedora_version = sorted_versions[0]
                fedora_release = fedora_rel_info
                return fedora_version, fedora_release # latest_rel from sorted_versions[0]
        else:
            fedora_rel_info = get_latest_href_from_pkg_url(f'{pkg_url}/{sorted_versions[0]}')
            fedora_version = sorted_versions[0]
            fedora_release = fedora_rel_info
            return fedora_version, fedora_release # latest_rel from sorted_versions[0]
   
    return (fedora_version, fedora_release)

def get_fedora_info_from_kojipkgs_pkgid(pkg):
    
    pkgid_url = f'{koji_pkgid_url}{pkg}'
    
    fedora_version = 'Not_Found'
    fedora_release_str = 'Not_Found'
    # Send a GET request to the URL
    response = requests.get(pkgid_url)

    # Check if the request was successful
    if response.status_code == 200:
        # Parse the HTML content
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Find the section with id 'taglist'
        taglist_section = soup.find(id='taglist')

        # If the taglist section is found, find the first table before it
        if taglist_section:
            # Get the first table before the taglist section
            first_table = taglist_section.find_previous('table')
            
            # Check if the first table was found
            if first_table:
                # Initialize a list to hold the build data
                builds_list = []
                
                # Iterate through the rows of the first table
                for row in first_table.find_all('tr')[1:]:  # Skip the header row
                    cols = row.find_all('td')
                    if len(cols) >= 4:  # Ensure there are enough columns
                        name = cols[0].text.strip()
                        version = cols[1].text.strip()
                        release = cols[2].text.strip()

                        # Determine the state from the class of the <td> element
                        state_cell = cols[3]
                        if 'complete' in state_cell['class']:
                            state = 'complete'
                        elif 'failed' in state_cell['class']:
                            state = 'failed'
                        else:
                            state = None  # Handle unexpected cases

                        # Append the data to the list if state is complete
                        if state == 'complete':
                            builds_list.append({
                                'Name': name,
                                'Version': version,
                                'Release': release,
                                'State': state
                            })
                
                # Filter for the desired names
                prioritized_names = [build['Name'] for build in builds_list if build['Name'].endswith('.fc41')]
                if prioritized_names:
                    result_name = prioritized_names[0]  # Get the first one if it exists
                else:
                    # Check for other versions
                    for version in range(40, 30, -1):  # Check .fc40 to .fc36
                        prioritized_names = [build['Name'] for build in builds_list if build['Name'].endswith(f'.fc{version}')]
                        if prioritized_names:
                            result_name = prioritized_names[0]  # Get the first one found
                            break
                    else:
                        result_name = None  # No matching names found

                # Output the result
                if result_name:
                    # Split the result name using '-' followed by a number as delimiter
                    split_name = re.split(r'-(?=\d)', result_name)
                    # print(f"Selected Name: {result_name}")
                    print(f"Split Name: {split_name}")
                    if split_name:
                        fedora_version = split_name[1]
                        fedora_release_str = split_name[2]
                else:
                    print("No matching Name found.")
            else:
                print("No table found before the taglist section.")
        else:
            print("No taglist section found on the page.")
    else:
        print(f"Failed to retrieve data. Status code: {response.status_code}")
    return (fedora_version, fedora_release_str)


def update_latest_fedora_pkg_info() -> None:
    """update details of the pkg from fedora
    """

    for pkg in pkg_list:
        if extract_fedora_info_from_kojipkgs_pkgid:
            fedora_version, fedora_release_str = get_fedora_info_from_kojipkgs_pkgid(pkg)
            logger.info(f'pkg : {pkg} fedora_version : {fedora_version} fedora_release_str : {fedora_release_str}')
            print(f'pkg : {pkg} fedora_version : {fedora_version} fedora_release_str : {fedora_release_str}') 
            if fedora_version != 'Not_Found'and '.fc41' not in fedora_release_str: # pkg version found but no fc41, needs revisit
                updatexl_pkg_col_value(pkg, 17, 'Not_Found')
                fedora_rel = fedora_release_str.split('.')[-1]
                msg = f'Needs Review: fedora_ver = {fedora_version} fedora_release = {fedora_rel}'
                updatexl_pkg_col_value(pkg, 21, msg, revisit_color) # col 'U' 
            else: # this will cover for both pkg fedora ver found / not found 
                updatexl_pkg_col_value(pkg, 17, fedora_version)       
        elif extract_fedora_info_from_kojipkgs:
            fedora_version, fedora_release_str = get_fedora_pkg_info_from_kojipkgs_url(pkg)
            logger.info(f'pkg : {pkg} fedora_version : {fedora_version} fedora_release_str : {fedora_release_str}')
            print(f'pkg : {pkg} fedora_version : {fedora_version} fedora_release_str : {fedora_release_str}') 
            if fedora_version != 'Not_Found'and '.fc41' not in fedora_release_str: # pkg version found but no fc41, needs revisit
                updatexl_pkg_col_value(pkg, 17, 'Not_Found')
                fedora_rel = fedora_release_str.split('.')[-1]
                msg = f'Needs Review: fedora_ver = {fedora_version} fedora_release = {fedora_rel}'
                updatexl_pkg_col_value(pkg, 21, msg, revisit_color) # col 'U' 
            else: # this will cover for both pkg fedora ver found / not found 
                updatexl_pkg_col_value(pkg, 17, fedora_version)
                
        else:
            fedora_version, fedora_release_str, upstream_src = get_fedora_pkg_info_from_pkg_url(pkg)
            if fedora_version == 'Not_Found':
                logger.error(f'{pkg}:info not found in pkg_url, Searching the pkg info from git')
                fedora_version, fedora_release_str, upstream_src = get_fedora_pkg_info_from_git(pkg)
            if 0: #fedora_release_str != None and cur_stable_fedora_rel not in fedora_release_str and '.fc' in fedora_release_str and fedora_version != 'Not_Found':# update Q, R as Not_Found and write the contents of Q & R in U
                updatexl_pkg_col_value(pkg, 17, 'Not_Found') # col 'Q'          
                updatexl_pkg_col_value(pkg, 20, 'Not_Found') # col 'T' 
                fedora_rel = fedora_release_str.split('.')[-1]
                msg = f'Needs Review: fedora_ver = {fedora_version} fedora_release = {fedora_rel} upstream_src = {upstream_src}'
                updatexl_pkg_col_value(pkg, 21, msg, revisit_color) # col 'U'             
            else: 
                # Sometimes pkg versions has chars like '~' example libtommath
                # Sometimes pkg versions has chars like '^' example libusbmuxd
                # Lets discard the string after encountering any special chars
                fedora_version_stripped = discard_after_special_chars(fedora_version, special_chars)
                if fedora_version_stripped == fedora_version : # No change
                    updatexl_pkg_col_value(pkg, 17, fedora_version_stripped) # col 'Q'
                else:
                    updatexl_pkg_col_value(pkg, 17, fedora_version_stripped, revisit_color) # col 'Q'
                    msg = f'Needs Review:  fedora_ver = {fedora_version}'
                    updatexl_pkg_col_value(pkg, 21, msg, revisit_color) # Update Col U about the weird fedora version          
                updatexl_pkg_col_value(pkg, 20, upstream_src) # col 'T'
        
        if fedora_version != 'Not_Found' and fedora_release_str != 'Not_Found':
            updatexl_fedora_src_link(pkg, fedora_version, fedora_release_str) # update col 'R'
        

def update_current_pkg_versions():
    """Update current version of packages found in spec files of 2.0 & 3.0
    """
    pkg_ver_2_0 = ''
    pkg_ver_3_0 = ''
    for pkg in pkg_list:
        pkg_ver_2_0 = get_pkg_ver(pkg, '2_0')
        pkg_ver_3_0 = get_pkg_ver(pkg, '3_0')
        updatexl_pkg_col_value(pkg, 15, pkg_ver_2_0) # col 'O'
        if 'Moved_to_Core' in pkg_ver_3_0 :
            updatexl_pkg_col_value(pkg, 16, 'Not_Found') # col 'P'
            updatexl_pkg_col_value(pkg, 21, pkg_ver_3_0, revisit_color) # col 'U'
        else:
            updatexl_pkg_col_value(pkg, 16, pkg_ver_3_0) # col 'P'


def update_daily_build_status() -> None:
    """Extract build status and date from the build_state.csv and update XL
    """
    cols = ['Package','State']
    table = pd.read_csv(build_state, usecols=cols)
    for pkg in pkg_list:
        pkg_state = 'Not_Found'        
        for row in table.itertuples():
            # bs_pkg_str = 
            bs_pkg_name = re.split(r"-(?=\d)", row[1])

            if pkg == bs_pkg_name[0]:
                logger.info(f'bs_pkg_name:{bs_pkg_name[0]}')
                print(f'bs_pkg_name:{bs_pkg_name[0]}')
                pkg_state = row[2]
                break
        logger.info(f'{pkg}:{pkg_state}')
        updatexl_build_status_and_date(pkg, pkg_state)


def update_pkg_status():
    """ For every package compare the values from "Need Upgrade",col D 
        and "Daily build status", col M and update pkg "Status", col H
    """
    pkg_status = ['Not_Started', 'Ongoing', 'PR Raised', 'PR in Review', 'Done-Upgrade', 'Done-FixBuild', 'Done-OtherChanges', 'NA-Uptodate']
    # Load the workbook
    wb = openpyxl.load_workbook(workbook)  
    # Access the worksheet (assuming it's the first sheet) 
    ws = wb.active
    for pkg in pkg_list:
        need_upgrade = ''
        build_status = ''    
        for r in ws.iter_rows(min_row=3, values_only=True):
            if r[2] != None:
                if r[2].strip(' \t\n\r') == pkg: #Skip row 0 & 1 , pkg info starts from row 2
                    need_upgrade = ws.cell(row=r[1]+2, column=4).value # col 'D'
                    build_status = ws.cell(row=r[1]+2, column=13).value # col 'M'
                    if build_status == 'Y' and need_upgrade == 'Y':
                        updatexl_pkg_col_value(pkg, 8, pkg_status[0]) # col 'H'
                    elif build_status == 'Y' and need_upgrade == 'N':
                        updatexl_pkg_col_value(pkg, 8, pkg_status[7]) # col 'H'
                    elif build_status == 'N' and need_upgrade == 'N':
                        updatexl_pkg_col_value(pkg, 8, pkg_status[0]) # col 'H'
                    else:
                        updatexl_pkg_col_value(pkg, 8, pkg_status[0]) # col 'H'    
        

def submit_entries():
    """Accepts input paths from the user and validates
    """
    values = []
    global workbook, build_state, code_dir_3_0, code_dir_2_0
    for entry in entries:
        values.append(entry.get())
    # print(values)
    for value in values:
        if not Path(value).exists():
            logger.info(f'Path: {value} does not exist')
            messagebox.showerror(title=f'Path Error:', message=f'{value} does not exist')
            
    workbook = f'{values[0]}'
    build_state = f'{values[1]}'
    code_dir_3_0 = f'{values[2]}'
    code_dir_2_0 = f'{values[3]}'
    print(f'Paths entered\n workbook = {workbook}\n build_state = {build_state}\n code_dir_3_0 = {code_dir_3_0}\n code_dir_2_0 = {code_dir_2_0}')
    logger.info(f'Paths entered\n workbook = {workbook}\n build_state = {build_state}\n code_dir_3_0 = {code_dir_3_0}\n code_dir_2_0 = {code_dir_2_0}')
    

    

def get_file_paths_from_user():
    """ Accepts various file paths as input from the user 
    """
    global workbook
    window = tk.Tk()
    window.title(f"Note: Enter Paths with \\\\ as shown in the examples")
    window.geometry('{}x{}'.format(600, 300))

    input_msgs = ['Excel Sheet File Path', 'Build State File Path', '3.0 Code directory', '2.0 Code Directory']
    input_path_examples = ["e.g. C:\\\\work\\\\PackageUpgradeReport.xlsx", 'e.g. C:\\\\work\\\\pkg_artifacts\\\\build_state.csv', 'e.g. C:\\\\work\\\\al3_0\\\\azurelinux', 'e.g. C:\\\\work\\\\al2_0\\\\azurelinux']
    for i in range(4):
        label = tk.Label(window, text=f"{input_msgs[i]}:")
        label.grid(row=i, column=0)

        entry = tk.Entry(window)
        entry.grid(row=i, column=1)
        
        label = tk.Label(window, text=f"{input_path_examples[i]}")
        label.grid(row=i, column=2)
        
        entries.append(entry)

    submit_button = tk.Button(window, text="Submit", command=submit_entries)
    submit_button.grid(row=4, column=0, columnspan=2,)

    msg = f"Please close XL file {workbook} and then click continue"
    msg_button = tk.Message(window, text=msg)
    msg_button.grid(row=6, column=0, columnspan=2)
    exit_button = tk.Button (window, text="Continue.", command = window.destroy)
    exit_button.grid(row=7, column=0, columnspan=2)

    window.mainloop()    

                
def calculate_time(func: Callable[[], None]) -> Callable[[], None]:
    """Decorator to calculate duration taken by any function.

    Args:
        func (Callable[[], None]): A callable function to be wrapped.

    Returns:
        Callable[[], None]: The wrapped callable function.
    """
    @wraps(func)
    def inner1() -> None:

        # storing time before function execution
        begin = time.time()

        func()

        # storing time after function execution
        end = time.time()
        print("Total time taken in second(s): ", func.__name__, end - begin)

    return inner1



def extract_src_rpm(rpm_file, output_dir):
    """Extracts the contents of an src.rpm file using 7-Zip.

    Args:
        rpm_file (_type_): _description_
        output_dir (_type_): _description_
    """

    cmd = [
        "7z",
        "x",  # Extract command
        rpm_file,
        "-o" + output_dir,  # Output directory
    ]

    subprocess.run(cmd, check=True)


def cleanup_xl_sheet():
    """ Remove spaces from around pkg names
        Remove version number
        Delete empty rows
    """
    wb = openpyxl.load_workbook(workbook)  
    ws = wb.active
    # Iterate over rows and delete based on a condition

    for row in range(ws.max_row+1, 1, -1):  ##range is from bottom to top, step -1 
        if ws[row][1].value is None:
            ws.delete_rows(idx=row, amount = 1)
    wb.save(workbook)
    
    for row in ws.iter_rows(min_row=3, values_only=True): #min_row=3 as pkg names start from row 3        
        if row[2] != None:
            pkg = row[2].strip(' \t\n\r')
            # pkg_list.append(pkg) # row[2] = our col 'C'
            pkg = re.split(r"-(?=\d)", pkg)
            ws.cell(row=row[1]+2, column=3).value = pkg[0]
    wb.save(workbook)
    
# Main function
@calculate_time
def main() -> None:
    
    print('Please wait for the app to load ....')
    get_file_paths_from_user()
    # pkg = 'kyotocabinet'
    # version = '1.2.80'
    # release = '6.fc41'
    # url = f'https://kojipkgs.fedoraproject.org//packages/{pkg}/{version}/{release}/src/{pkg}-{version}-{release}.src.rpm'
    # urllib.request.urlretrieve(url, f'{pkg}.src.rpm')
    # extract_src_rpm(f'{pkg}.src.rpm', f'{pkg}')
    print('Now Processing ........Please wait')
    print('For more info you may refer the log file: pkg_update_analysis.log')
    cleanup_xl_sheet()
    print('Now Processing: read_all_pkg_names. Please wait.........')
    logger.info('Now Processing: read_all_pkg_names. Please wait.........')
    read_all_pkg_names()
    print('Now Processing: update_daily_build_status. Please wait.........')
    logger.info('Now Processing: update_daily_build_status. Please wait.........')
    update_daily_build_status()
    print('Now Processing: update_current_pkg_versions. Please wait.........')
    logger.info('Now Processing: update_current_pkg_versions. Please wait.........')
    update_current_pkg_versions()
    print('Now Processing: update_latest_fedora_pkg_info. Please wait.........')
    logger.info('Now Processing: update_latest_fedora_pkg_info. Please wait.........')
    update_latest_fedora_pkg_info()
    print('Now Processing: update_if_need_upgrade. Please wait.........')
    logger.info('Now Processing: update_if_need_upgrade. Please wait.........')
    update_if_need_upgrade()
    print('Now Processing: update_pkg_status. Please wait.........')
    logger.info('Now Processing: update_pkg_status. Please wait.........')
    update_pkg_status()
    print('All Processing: DONE')
    logger.info('All Processing: DONE')
    while True:
        user_input = input("Type 'x' to quit): ")
        if user_input == "x":
            sys.exit("Exiting program...")
        else:
            print("You entered:", user_input)

if __name__ == '__main__':
    main()